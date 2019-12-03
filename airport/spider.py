# -*- coding:utf-8 -*-
import os
import re
import sys
import time
import queue
import platform
from threading import Thread
from geopy.geocoders import Nominatim
from typing import List, Dict, NewType, Tuple, Optional
from selenium import webdriver
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.remote.webelement import WebElement
from openpyxl import Workbook, load_workbook

Country = NewType("Country", Tuple[str, str])
Pagination = NewType("Pagination", Tuple[int, str])
Airport = NewType("Airport", Tuple[str, str, str, str, str, str])
AirportDetail = NewType("AirportDetail", Tuple[str, str, str, str, str, str,
                                               str, str])
start_url = "http://airport.anseo.cn/"

country_pattern = r"(\w+)\((.{2,})\)"

process__queue = queue.Queue()
is_finishing = False
xls_file_path = "./机场.xlsx"


def start_crawl(**settings):
    country_names: List[str] = settings["COUNTRY"]
    columns: List[str] = settings["COLUMN"]
    crawl_task = __init_crawl_task(country_names)
    process_task = __start_process_task(country_names, columns)
    crawl_task.start()
    process_task.start()
    process_task.join()
    print("爬虫执行结束")


def __init_crawl_task(country_names: List[str]) -> Thread:
    return Thread(target=__init_crawl_job,
                  name="airport_crawl_task_0",
                  args=(country_names, ))


def __init_crawl_job(country_names: List[str]) -> None:
    chrome_driver = None
    try:
        print("start crawl data ......")
        chrome_driver = __init_chrome_driver()
        chrome_driver.get(start_url)
        country_list: List[Country] = __parse_country_list(
            chrome_driver, country_names)
        gps = Nominatim(user_agent="airport_spider", timeout=30)
        for country in country_list:
            (country_name, country_href) = country
            print(f"[{country_name}]机场列表链接:{country_href}")
            # 访问具体国家机场列表页
            chrome_driver.get(country_href)
            airport_page_list = __parse_airport_page_list(chrome_driver)
            print(f"[{country_name}]共有[{len(airport_page_list)}]页机场信息")
            for airport_page in airport_page_list:
                (page_index, page_href) = airport_page
                # 访问机场列表页面
                print(f"访问[{country_name}]机场列表第[{page_index}]页:{page_href}")
                chrome_driver.get(page_href)
                airport_list = __parse_airport_list(chrome_driver,
                                                    country_name)
                for airport in airport_list:
                    chrome_driver.get(airport[3])
                    detail_tuple = __parse_airport_detail(chrome_driver)
                    airport_detail = airport + detail_tuple
                    addr = ",".join(airport_detail[0:3])
                    chrome_driver.get("https://maplocation.sjfkai.com/")
                    chrome_driver.find_element_by_xpath(
                        '//*[@id="locations"]').send_keys(addr)
                    chrome_driver.find_element_by_xpath(
                        '//*[@id="platform"]/div[2]/label/span[1]/input'
                    ).click()
                    chrome_driver.find_element_by_xpath(
                        '//*[@id="root"]/div/div/div[2]/form/div/div[2]/div/div[3]/div/div/div/span/button'
                    ).click()
                    time.sleep(5)
                    lat = chrome_driver.find_element_by_xpath(
                        '//tbody/tr[1]/td[4]').text
                    lng = chrome_driver.find_element_by_xpath(
                        '//tbody/tr[1]/td[3]').text
                    # lng_lat = point.split(",")
                    airport_detail = airport_detail + (lat, lng)
                    process__queue.put(airport_detail, False, 30)
                    #airport_detail_list.append(airport_detail)
                    print(airport_detail)
    except:
        print("爬虫执行异常=", sys.exc_info())
    finally:
        global is_finishing
        is_finishing = True
        if chrome_driver is not None:
            chrome_driver.quit()


def __init_chrome_driver() -> WebDriver:
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    driver_path = "chromedriver.exe"
    if platform.system() == "Darwin":
        driver_path = "/Users/Nick/Dev/Python-In-Action/airport/chromedriver"
    chrome_driver = webdriver.Chrome(executable_path=driver_path,
                                     options=chrome_options)
    chrome_driver.set_page_load_timeout(90)
    chrome_driver.set_script_timeout(5)
    chrome_driver.implicitly_wait(5)
    return chrome_driver


def __parse_country_list(chrome_driver: WebDriver,
                         country_names: List[str]) -> List[Country]:
    all_country_el_list = chrome_driver.find_elements_by_xpath(
        '//div[@class="mod-body"]//a')
    country_el_list = list(
        filter(
            lambda el: re.match(country_pattern, el.get_attribute("title")) !=
            None, all_country_el_list))
    country_el_list = list(
        filter(
            lambda el: re.match(country_pattern, el.get_attribute("title")).
            group(2) in country_names, country_el_list))
    country_list = list(
        map(
            lambda el:
            (re.match(country_pattern, el.get_attribute("title")).group(2),
             el.get_attribute("href")), country_el_list))
    return country_list


def __parse_airport_page_list(chrome_driver: WebDriver) -> List[Pagination]:
    airport_page_list: List[Pagination] = []
    try:
        pagination_el = chrome_driver.find_element_by_xpath(
            '//ul[@class="pagination pull-right"]')
        a_el_list = pagination_el.find_elements_by_xpath("//li/a")
        a_el_list = list(
            filter(
                lambda el: re.search("__page-(\d{1,})", el.get_attribute(
                    "href")) != None, a_el_list))
        page_max = max(
            list(
                map(
                    lambda el: int(
                        el.get_attribute("href").split("__page-")[1]),
                    a_el_list)))
        a_el_href: str = a_el_list[-1].get_attribute("href")
        href_prefix = a_el_href.split("__page-")[0]
        for page_index in range(1, page_max + 1):
            airport = (page_index, href_prefix + f"__page-{page_index}")
            airport_page_list.append(airport)
    except:
        airport_page_list.append((1, chrome_driver.current_url))
    return airport_page_list


def __parse_airport_list(chrome_driver: WebDriver,
                         country_name: List[str]) -> List[Airport]:
    airport_list: List[Airport] = []

    row_el_list = chrome_driver.find_elements_by_xpath("//table/tbody/tr")
    for row_el in row_el_list:
        city_name = None
        airport_name = None
        try:
            city_name = row_el.find_element_by_xpath("./td[1]/a").text.split(
                "\n")[-1]
        except:
            city_name = row_el.find_element_by_xpath(
                "./td[1]/a/font").text.split("\n")[-1]
        try:
            airport_name = row_el.find_element_by_xpath(
                "./td[2]/a").text.split("\n")[-1]
        except:
            airport_name = row_el.find_element_by_xpath(
                "./td[2]/a/font").text.split("\n")[-1]
        airport_href = row_el.find_element_by_xpath("./td[2]/a").get_attribute(
            "href")
        code_3 = row_el.find_element_by_xpath("./td[3]//a").text
        code_4_split = row_el.find_element_by_xpath(
            "./td[4]/span").get_attribute("title").split(":")
        code_4 = code_4_split[-1]
        airport = (country_name, city_name, airport_name, airport_href, code_3,
                   code_4)
        airport_list.append(airport)
    return airport_list


def __parse_airport_detail(chrome_driver: WebDriver) -> Tuple:
    phone = chrome_driver.find_element_by_xpath(
        '//ul[@class="info-detail"]/li[5]').text.split("：")[-1]
    description = None
    try:
        description = chrome_driver.find_element_by_xpath(
            '//div[@class="airport-des-c"]/p').text
    except:
        description = ""
    return (phone, description)


def __init_xls(country_list: List[str], columns: List[str]) -> Workbook:
    print("正在初始化表格...")
    workbook = None
    if os.path.exists(xls_file_path):
        workbook = load_workbook(xls_file_path)
    else:
        workbook = Workbook()
        workbook.remove(workbook.get_active_sheet())
        for country_name in country_list:
            worksheet = workbook.create_sheet(country_name)
            worksheet.sheet_properties.tabColor = '6FB7B7'
            for xls_col in worksheet.iter_cols(max_col=len(columns),
                                               max_row=1):
                for cell in xls_col:
                    cell.value = columns[cell.col_idx - 1]
    return workbook


def __write_xls(workbook: Workbook, position) -> None:
    print("正在写入数据...")
    worksheet = workbook.get_sheet_by_name(position[0])
    worksheet.append(position[1:])


def __close_xls(workbook: Workbook, filename: str) -> None:
    workbook.save(filename)
    print("数据文件已保存.")


def __start_process_task(country_list: List[str],
                         columns: List[str]) -> Thread:
    print("正在运行数据处理任务...")
    workbook = __init_xls(country_list, columns)
    thread = Thread(target=__process_position,
                    name="position_task_0",
                    args=(workbook, ))
    return thread


def __process_position(workbook: Workbook) -> None:
    while not is_finishing:
        while not process__queue.empty():
            position = process__queue.get()
            __write_xls(workbook, position)
    __close_xls(workbook, xls_file_path)
